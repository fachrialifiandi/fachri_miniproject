from openpyxl import load_workbook

excel_file_name = 'FinancialRecord.xlsx'
indoesian_rupiah = '"Rp."#,##0'
max_row = 32


try:

    wb = load_workbook(excel_file_name)
    ws = wb.active
    print(f"Successfully to detect {excel_file_name}")

    total_column_index = None

    # Find out in which column the total is
    for cell in ws[1]:
        if cell.value and 'total' in str(cell.value).lower():
            total_column_index = cell.column
            break

    # If the total is not found,
    # use max_column as a safe limit
    if total_column_index is None:
        total_column_index = ws.max_column

    stop = False
    # Use the loop to fill in data
    for row_index in range(2, max_row):
        print(f'\n --- Data fill for row {row_index} --- ')

        # daily_expenses = []

        for col_index in range(1, total_column_index):

            header = ws.cell(row=1, column=col_index).value

            user_input = input(f"Input '{header}' : ")

            if user_input.lower() == 'stop':
                stop = True
                break

            elif user_input.lower() == 'next':
                break

            cell = ws.cell(row=row_index, column=col_index)

            try:
                expense_value = int(user_input)
                cell.value = expense_value
                cell.number_format = indoesian_rupiah

            except (ValueError, TypeError):
                cell.value = user_input

        # if not stop and user_input != 'next':
        #     total_today = sum(daily_expenses)
        #     print(f'\nTotal expenses today : {total_today}')

        if stop:
            print('\n STOP command received. ')
            break

    wb.save(excel_file_name)
    print(f"✔ Data has been save to {excel_file_name}")


except FileNotFoundError:
    print(f"Error file with name : '{excel_file_name}' is not found. ")

except Exception as e:
    print(f'Error Found : {e}')
