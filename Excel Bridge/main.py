from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# wb = Workbook()  # intialize
wb = load_workbook('.xlsx')
ws = wb.active


# Move and Copy Cells-
ws.move_range("C1:D11", rows=1, cols=2)
wb.save('.xlsx')







# Delete and Insert Rows
# ws.delete_rows(7)
# ws.insert_rows(7)
# ws.insert_cols(2)
# ws.delete_cols(2)


# Merge and Unmerge
# ws.merge_cells("A1:D1")
# ws.unmerge_cells("A1:D1")


# Accessing multiple cells
# for row in range(1, 11):
#     for col in range(1, 5):
#         char = get_column_letter(col)
#         ws[char + str(row)] = "Tomiyasu"
# print(ws[char + str(row)].value)


# Adding Rows
# ws.append(['Maybe', 'Terrible', 'Yesterday', '?'])
# ws.append(['Love', 'Story', 'Key', '?'])
# ws.append(['Now', 'Better', 'Rest', ':)'])


# ws = wb.active
# ws['A2'].value = "Test"
# wb.save('grades.xlsx')


# wb.create_sheet("Test")
# print(ws['A1'].value)
# print(wb.sheetnames)
