from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
    "Tika": {
        "Math": 80,
        "Science": 56,
        "English": 45,
        "Historical": 76

    },
    "Gemma": {
        "Math": 76,
        "Science": 67,
        "English": 76,
        "Historical": 45

    },
    "Tara": {
        "Math": 98,
        "Science": 88,
        "English": 65,
        "Historical": 70
    },

    "Jason": {
        "Math": 86,
        "Science": 66,
        "English": 43,
        "Historical": 95
    },

    "Kevin": {
        "Math": 21,
        "Science": 91,
        "English": 78,
        "Historical": 78

    }

}


wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ['Name'] + list(data['Tika'].keys())
ws.append(headings)

for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

for col in range(2, len(data["Tika"]) + 2):
    char = get_column_letter(col)
    ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"
    
for col in range(1,6):
    ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")
    
    

wb.save("NewGrades.xlsx")
